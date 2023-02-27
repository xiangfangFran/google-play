import os
import re
import json
import time
import pymysql
import requests
import multiprocessing
from loguru import logger
from openpyxl import Workbook
from datetime import datetime
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed


PROXY = {
    "http": "http://127.0.0.1:10809",
    "https": "http://127.0.0.1:10809"
}
# PROXY= {
# }


def save_excel_file(app_result):
    sheet_name = "Sheet"
    if app_result[0]['brand_id'] == 1:
        sheet_name = "免费榜"
    elif app_result[0]['brand_id'] == 5:
        sheet_name = "人气飙升榜"
    wb = Workbook()
    ws = wb.create_sheet(sheet_name)
    titles = ['排行榜', 'APP名称', '游戏类别', '下载量', '发布日期', '更新日期', 'GP链接', '应用简介', '简介翻译', '应用描述',
              '应用描述翻译', 'LOGO', 'LOGO主色调', 'GP展示图']
    for i in range(len(titles)):
        ws.cell(row=1, column=i+1, value=titles[i])
    row = 2
    for app in app_result:
        ws.cell(row=row, column=2, value=app['genre_ranking'])
        ws.cell(row=row, column=3, value=app['name'])
        # ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value=app['download_times'])
        ws.cell(row=row, column=6, value=app['release_time'])
        ws.cell(row=row, column=7, value=app['last_release_time'])
        ws.cell(row=row, column=8, value=app['app_url'])
        ws.cell(row=row, column=9, value=app['title'])
        ws.cell(row=row, column=11, value=app['content'])
        ws.cell(row=row, column=13, value=app['logo'])
        ws.cell(row=row, column=13, value=app['colors'])
        ws.cell(row=row, column=13, value="见之前文件")
        row = row+1
    wb.save("/data/IND-apps.xlsx")
    

def app_colors_process(bundle_id):
    text = None
    url = f"https://api.diandian.com/pc/app/v1/competitor/app_info_comparison?market_id={market_id}&app_ids={bundle_id}&country_id={country_id}&language_id={language_id}&k={key}"
    response = requests.get(url=url).json()
    if response['code'] == 0 and response['message'] == "success":
        colors = response['data']["list"][0]["color_distribution"]
        text = ""
        for color in colors:
            text = text + f"{color[0]}-{color[1]}%;"
    return text
        

def detail_app_process(detail_url, rank, referer):
    start_time = time.time()
    logger.info(f'开始获取商品详情页, detail_url: {detail_url}')
    # url = f"https://app.diandian.com{detail_url}"
    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "cookie": 'deviceid=0e28f41bb884cc3a4790e5693557c14; token=e463bd6c0419b7c6238c5f4d94fe1264d9e42a56ee30af02ecf76313d9cee25acffc1cd0503aac86b5435e5bd97aa481749cea90b8466f7e7107f7907169acbf37ed489371461baee7497bad0429cfe3; Qs_lvt_404253=1677118072%2C1677148704%2C1677211226; Qs_pv_404253=2707789470116610000%2C4364250506636174000',
        "referer": referer
    }
    response = requests.get(url=detail_url, headers=headers, proxies=PROXY).text
    if "Server error" in response:
        # raise Exception("Server error")
        logger.error("Server error")
        return "", ""
    logger.success(f'获取商品详情页完成, detail_url: {detail_url}, 耗时: {time.time()-start_time}')
    soup = BeautifulSoup(response, "html.parser")
    name = soup.find("div", class_="dd-flex dd-align-center dd-overflow-hidden").get_text(strip=True)
    err = ["?",",","_","/","*",""",""","<",">","|"]
    for i in err:
        name = name.replace(i, "")
    logo_div = soup.find("a", class_="logo-wrap")
    if logo_div:
        logo = logo_div.find("img", class_="dd-app-logo")
        start_save_logo_time = time.time()
        logger.info(f'开始下载logo, name: {rank}-{name}.jpg')
        logo_data = requests.get(url=logo["src"]).content
        with open(f"data/logo/{rank}-{name}.jpg", "wb") as fp:
            fp.write(logo_data)
        logger.success(f'下载logo完成, name: {rank}-{name}.jpg, 耗时: {time.time()-start_save_logo_time}')
    backgroud_imgs = soup.find_all("img", class_="el-image screen-img")
    for i in range(len(backgroud_imgs)):
        start_save_screen_time = time.time()
        logger.info(f'开始下载应用截图, name: {rank}-{name}-{i+1}.jpg')
        img_data = requests.get(url=backgroud_imgs[i]['src']).content
        with open(f"data/背景图/{rank}-{name}-{i+1}.jpg", "wb") as fp:
            fp.write(img_data)
        logger.success(f'下载应用截图完成, name: {rank}-{name}-{i+1}.jpg, 耗时: {time.time()-start_save_screen_time}')
    content = soup.find("p", class_="main-content")
    content = content.get_text() if content else None
    download_times = None
    tag = soup.find("div", class_="app-info-card-item")
    if "累计下载量" in tag.get_text():
        download_times = soup.find("a", class_="app-value dd-flex dd-flex-center dd-align-center dd-base-color font-16 font-500")
        download_times = download_times.get_text(strip=True)
    return content, download_times


def get_brand_page(brand_id):
    results = []
    today = f'{datetime.strftime(datetime.now(), "%Y-%m-%d")} 00:00:00'
    today_time_stamp = int(time.mktime(time.strptime(today, "%Y-%m-%d %H:%M:%S")))*1000
    headers = {
        "Host": "api.diandian.com",
        "Origin": "https://app.diandian.com",
        "Referer": "https://app.diandian.com/",
        "Accept": "application / json, text / plain, * / *",
        "cookie": 'deviceid=0e28f41bb884cc3a4790e5693557c14; Qs_lvt_404253=1677118072; Qs_pv_404253=1760908230870957800%2C1254500648111455500%2C4565208287553404400%2C2698375595445380000%2C3324556445801356300; token=e463bd6c0419b7c6238c5f4d94fe1264d9e42a56ee30af02ecf76313d9cee25acffc1cd0503aac86b5435e5bd97aa481749cea90b8466f7e7107f7907169acbf37ed489371461baee7497bad0429cfe3'
    }
    url = f"https://app.diandian.com/rank/googleplay/{market_id}-{rank_type}-{genre_id}-{country_id}-{brand_id}?time={today_time_stamp}"
    response = requests.get(url=url, proxies=PROXY).text
    if "404 Page Not Find" in response:
        raise Exception("404 Page Not Find")
    soup = BeautifulSoup(response, "html.parser")
    div_tags = soup.find_all("div", class_="el-row is-align-middle el-row--flex border dd-hover-row")
    for div in div_tags:
        result = {}
        column_tags = div.find_all("div", class_=re.compile("el-col el-col-*"))
        for column_tag in column_tags:
            if column_tag.find("div", class_="dd-app right-info"):
                head_tag = column_tag.find("div", class_="dd-app right-info")
                app_a_tag = head_tag.find("a", class_="logo-img")
                result['app_url'] = "https://app.diandian.com"+app_a_tag['href']
                result['name'] = head_tag.find("div", class_="show-text dd-max-ellipsis").get_text()
                result['developer_name'] = head_tag.find("p", class_="font-12 dd-desc-font-color develop-info").get_text()
                day_tag = head_tag.find("span", class_="day-tag").get_text()
                day = re.search("\d+", day_tag)
                result['hegemony_days'] = day.group() if day else None
            elif column_tag.find("div", class_="total-rank"):
                root_rank_tags = column_tag.find("div", class_="dd-flex dd-align-center dd-flex-start position-relative")
                range_tag = root_rank_tags.find("div", class_=re.compile("range.*"))
                rank_tags = root_rank_tags.find("div", class_="total-rank")
                for rank_tag in rank_tags.find_all("div"):
                    if rank_tag.get_text() == "总榜":
                        result['brand_ranking'] = rank_tags.get_text(strip=True).replace("总榜", "")
                    elif rank_tag.get_text() == "游戏":
                        result['category_ranking'] = rank_tags.get_text(strip=True).replace("游戏", "")
                        if range_tag.find("i", class_="iconfont Dianxiajiang"):
                            result['category_ranking_incr'] = f"-{range_tag.get_text(strip=True)}"
                        elif range_tag.find("i", class_="iconfont Dianshangsheng"):
                            result['category_ranking_incr'] = f"+{range_tag.get_text(strip=True)}"
                    elif rank_tag.get_text() == "赌场游戏":
                        result['genre_ranking'] = rank_tags.get_text(strip=True).replace("赌场游戏", "")
                        if range_tag.find("i", class_="iconfont Dianxiajiang"):
                            result['genre_ranking_incr'] = f"-{range_tag.get_text(strip=True)}"
                        elif range_tag.find("i", class_="iconfont Dianshangsheng"):
                            result['genre_ranking_incr'] = f"+{range_tag.get_text(strip=True)}"
            elif column_tag.find("div", class_="dd-text-center"):
                if column_tags.index(column_tag) == 5:
                    result['word_coverage'] = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 7:
                    result['rating_count'] = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 8:
                    result['release_time'] = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 9:
                    result['last_release_time'] = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
            elif column_tag.find("div", role="slider"):
                star_tag = column_tag.find("div", role="slider")
                result['rating'] = star_tag['aria-valuenow']
        content, download_times = detail_app_process(result['app_url'], result['genre_ranking'], url)
        bundle_id = re.search("/app/(.*?)/googleplay", result['app_url']).group(1)
        colors = app_colors_process(bundle_id)
        result['colors'] = colors
        result['content'] = content
        result['download_times'] = download_times
        result['brand_id'] = brand_id
    return results

if __name__ == "__main__":
    config_name = "config-local.json"

    with open(config_name, 'r') as fp:
        config = json.loads(fp.read())

    rank_url = config['task']['rank_url']
    app_url = config['task']['app_url']
    market_id = config['task']['market_id']
    genre_id = config['task']['genre_id']
    country_id = config['task']['country_id']
    language_id = config['task']['language_id']
    device_id = config['task']['device_id']
    rank_type = config['task']['rank_type']
    brand_list = config['task']['brand_id']
    key = config['task']['key']

    path = os.path.join(os.path.abspath("./"), "/data")
    if not os.path.exists(path):
        os.makedirs(path)
    status_file = os.path.join(path, "status.json")

    app_result = get_brand_page(brand_list[0])
    if app_result:
        save_excel_file(app_result)

