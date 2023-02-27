import os
import re
import json
import time
import pymysql
import requests
import multiprocessing
from loguru import logger
from openpyxl import Workbook
from datetime import datetime
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed


# PROXY = {
#     "http": "https://zjlt.missdaxiaojie2022.xyz:29775",
#     "https": "https://zjlt.missdaxiaojie2022.xyz:29775"
# }
PROXY= {
}


def save_excel_file(app_result):
    sheet_name = "Sheet"
    if app_result[0]['brand_id'] == 1:
        sheet_name = "免费榜"
    elif app_result[0]['brand_id'] == 5:
        sheet_name = "人气飙升榜"
    wb = Workbook()
    ws = wb.create_sheet(sheet_name)
    titles = ['排行榜', 'APP名称', '游戏类别', '下载量', '发布日期', '更新日期', 'GP链接', '应用简介', '简介翻译', '应用描述',
              '应用描述翻译', 'LOGO', 'LOGO主色调', 'GP展示图']
    for i in range(len(titles)):
        ws.cell(row=1, column=i+1, value=titles[i])
    row = 2
    for app in app_result:
        ws.cell(row=row, column=2, value=app['genre_raanking'])
        ws.cell(row=row, column=3, value=app['name'])
        # ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value=app['download_times'])
        ws.cell(row=row, column=6, value=app['release_time'])
        ws.cell(row=row, column=7, value=app['last_release_time'])
        ws.cell(row=row, column=8, value=f"https://play.google.com/store/apps/details?id={app['bundle_id']}&hl=hi&gl=in")
        ws.cell(row=row, column=9, value=app['title'])
        ws.cell(row=row, column=11, value=app['content'])
        ws.cell(row=row, column=13, value=app['logo'])
        ws.cell(row=row, column=13, value=app['colors'])
        ws.cell(row=row, column=13, value="见之前文件")
        row = row+1
    wb.save("/data/IND-apps.xlsx")
    

def app_colors_process(bundle_id):
    text = None
    url = f"https://api.diandian.com/pc/app/v1/competitor/app_info_comparison?market_id={market_id}&app_ids={bundle_id}&country_id={country_id}&language_id={language_id}&k={key}"
    response = requests.get(url=url).json()
    if response['code'] == 0 and response['message'] == "success":
        colors = response['data']["list"][0]["color_distribution"]
        text = ""
        for color in colors:
            text = text + f"{color[0]}-{color[1]}%;"
    return text
        

def detail_app_process(detail_url, rank):
# def detail_app_process(rank, bundle_id, app_id):
    start_time = time.time()
    # logger.info(f'开始获取商品详情页, rank: {rank}, bundle_id: {bundle_id}, app_id: {app_id}')
    logger.info(f'开始获取商品详情页, detail_url: {detail_url}')
    # url = f"https://app.diandian.com/app/{app_id}/googleplay?market={market_id}&country={country_id}"
    url = f"https://app.diandian.com{detail_url}"
    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "cookie": 'deviceid=0e28f41bb884cc3a4790e5693557c14; Hm_lvt_d185b2974609101d8f9340b5f861ca70=1676878943; Hm_lvt_8a5bd6e095cd118016489cab0443c2d7=1676878943; totalWatch=true; showMajorDialog=true; mediav=%7B%22eid%22%3A%221089729%22%2C%22ep%22%3A%22%22%2C%22vid%22%3A%22%25gpCBGJb3p9PDhUvJ%24%3Du%22%2C%22ctn%22%3A%22%22%2C%22vvid%22%3A%22%25gpCBGJb3p9PDhUvJ%24%3Du%22%2C%22_mvnf%22%3A1%2C%22_mvctn%22%3A0%2C%22_mvck%22%3A0%2C%22_refnf%22%3A0%7D; token=e463bd6c0419b7c6238c5f4d94fe1264d9e42a56ee30af02ecf76313d9cee25acffc1cd0503aac86b5435e5bd97aa481749cea90b8466f7e7107f7907169acbf37ed489371461baee7497bad0429cfe3; Qs_lvt_404253=1677118072%2C1677148704; Hm_lpvt_8a5bd6e095cd118016489cab0443c2d7=1677150528; Qs_pv_404253=2698375595445380000%2C3324556445801356300%2C4332587473021378000%2C4544993066206487000%2C1705137575899710700; Hm_lpvt_d185b2974609101d8f9340b5f861ca70=1677150528',
        "referer": "https://app.diandian.com/rank/googleplay/11-2-39-101-1?time=1677081600000"
    }
    response = requests.get(url=url, headers=headers).text
    if "Server error" in response:
        # raise Exception("Server error")
        logger.error("Server error")
        return ""
    # logger.success(f'获取商品详情页完成, rank: {rank}, bundle_id: {bundle_id}, app_id: {app_id}, 耗时: {time.time()-start_time}')
    logger.success(f'获取商品详情页完成, detail_url: {detail_url}, 耗时: {time.time()-start_time}')
    soup = BeautifulSoup(response, "html.parser")
    name = soup.find("div", class_="dd-flex dd-align-center dd-overflow-hidden").get_text(strip=True)
    err = ["?",",","_","/","*",""",""","<",">","|"]
    for i in err:
        name = name.replace(i, "")
    # hot_num = soup.find("h2", class_="btn-box-inner").get_text() if soup.find("h2", class_="btn-box-inner") else None
    logo_div = soup.find("a", class_="logo-wrap")
    if logo_div:
        logo = logo_div.find("img", class_="dd-app-logo")
        start_save_logo_time = time.time()
        logger.info(f'开始下载logo, name: {rank}-{name}.jpg')
        logo_data = requests.get(url=logo["src"]).content
        with open(f"data/logo/{rank}-{name}.jpg", "wb") as fp:
            fp.write(logo_data)
        logger.success(f'下载logo完成, name: {rank}-{name}.jpg, 耗时: {time.time()-start_save_logo_time}')
    backgroud_imgs = soup.find_all("img", class_="el-image screen-img")
    for i in range(len(backgroud_imgs)):
        start_save_screen_time = time.time()
        logger.info(f'开始下载应用截图, name: {rank}-{name}-{i+1}.jpg')
        img_data = requests.get(url=backgroud_imgs[i]['src']).content
        with open(f"data/背景图/{rank}-{name}-{i+1}.jpg", "wb") as fp:
            fp.write(img_data)
        logger.success(f'下载应用截图完成, name: {rank}-{name}-{i+1}.jpg, 耗时: {time.time()-start_save_screen_time}')
    content = soup.find("p", class_="main-content")
    content = content.get_text() if content else None
    download_times = None
    tag = soup.find("div", class_="app-info-card-item")
    if "累计下载量" in tag.get_text():
        download_times = soup.find("a", class_="app-value dd-flex dd-flex-center dd-align-center dd-base-color font-16 font-500")
        download_times = download_times.get_text(strip=True)
    return content, download_times
    

def get_brand_page(brand_id):
    today = f'{datetime.strftime(datetime.now(), "%Y-%m-%d")} 00:00:00'
    today_time_stamp = int(time.mktime(time.strptime(today, "%Y-%m-%d %H:%M:%S")))
    headers = {
        "Host": "api.diandian.com",
        "Origin": "https://app.diandian.com",
        "Referer": "https://app.diandian.com/",
        "Accept": "application / json, text / plain, * / *",
        "cookie": 'deviceid=0e28f41bb884cc3a4790e5693557c14; Qs_lvt_404253=1677118072; Qs_pv_404253=1760908230870957800%2C1254500648111455500%2C4565208287553404400%2C2698375595445380000%2C3324556445801356300; token=e463bd6c0419b7c6238c5f4d94fe1264d9e42a56ee30af02ecf76313d9cee25acffc1cd0503aac86b5435e5bd97aa481749cea90b8466f7e7107f7907169acbf37ed489371461baee7497bad0429cfe3'
    }
    url = f"https://app.diandian.com/rank/googleplay/{market_id}-{rank_type}-{genre_id}-{country_id}-{brand_id}?time={today_time_stamp}"
    response = requests.get(url=url, headers=headers).text
    soup = BeautifulSoup(response, "html.parser")
    div_tags = soup.find_all("div", class_="el-row is-align-middle el-row--flex border dd-hover-row")
    for div in div_tags:
        column_tags = div.find_all("div", class_=re.compile("el-col el-col-*"))
        for column_tag in column_tags:
            if column_tag.find("div", class_="dd-app right-info"):
                head_tag = column_tag.find("div", class_="dd-app right-info")
                app_a_tag = head_tag.find("a", class_="logo-img")
                app_url = "https://app.diandian.com/"+app_a_tag['href']
                name = ""
                app_name = head_tag.find("div", class_="show-text dd-max-ellipsis").get_text()
                err = ["?", ",", "_", "/", "*", """,""", "<", ">", "|"]
                for i in err:
                    name = app_name.replace(i, "")
                developer_name = head_tag.find("p", class_="font-12 dd-desc-font-color develop-info").get_text()
                day_tag = head_tag.find("span", class_="day-tag").get_text()
                day = re.search("\d+", day_tag)
                hegemony_days = day.group() if day else None
            elif column_tag.find("div", class_="total-rank"):
                root_rank_tags = column_tag.find("div", class_="dd-flex dd-align-center dd-flex-start position-relative")
                range_tag = root_rank_tags.find("div", class_=re.compile("range.*"))
                rank_tags = root_rank_tags.find("div", class_="total-rank")
                for rank_tag in rank_tags.find_all("div"):
                    if rank_tag.get_text() == "总榜":
                        brand_rank = rank_tag.get_text().replace("总榜")
                    elif rank_tag.get_text() == "游戏":
                        category_rank = rank_tag.get_text().replace("游戏")
                        if range_tag.find("i", class_="iconfont Dianxiajiang"):
                            category_rank_incr = f"-{range_tag.get_text(strip=True)}"
                        elif range_tag.find("i", class_="iconfont Dianshangsheng"):
                            category_rank_incr = f"+{range_tag.get_text(strip=True)}"
                    elif rank_tag.get_text() == "赌场游戏":
                        genre_rank = rank_tag.get_text().replace("赌场游戏")
                        if range_tag.find("i", class_="iconfont Dianxiajiang"):
                            genre_rank_incr = f"-{range_tag.get_text(strip=True)}"
                        elif range_tag.find("i", class_="iconfont Dianshangsheng"):
                            genre_rank_incr = f"+{range_tag.get_text(strip=True)}"
            elif column_tag.find("div", class_="dd-text-center"):
                if column_tags.index(column_tag) == 5:
                    word_coverage = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 7:
                    rating_count = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 8:
                    release_time = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 9:
                    last_release_time = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
            elif column_tag.find("div", role="slider"):
                star_tag = column_tag.find("div", role="slider")
                rating = star_tag['aria-valuenow']
    
                
    

def list_app_process(brand_id):
    page = 1
    total_num = 0
    count = 0
    app_records = []
    developer_records = []
    today = f'{datetime.strftime(datetime.now(), "%Y-%m-%d")} 00:00:00'
    today_time_stamp = int(time.mktime(time.strptime(today, "%Y-%m-%d %H:%M:%S")))
    headers = {
        "Host": "api.diandian.com",
        "Origin": "https://app.diandian.com",
        "Referer": "https://app.diandian.com/",
        "Accept": "application / json, text / plain, * / *",
        "cookie": 'deviceid=0e28f41bb884cc3a4790e5693557c14; Qs_lvt_404253=1677118072; Qs_pv_404253=1760908230870957800%2C1254500648111455500%2C4565208287553404400%2C2698375595445380000%2C3324556445801356300; token=e463bd6c0419b7c6238c5f4d94fe1264d9e42a56ee30af02ecf76313d9cee25acffc1cd0503aac86b5435e5bd97aa481749cea90b8466f7e7107f7907169acbf37ed489371461baee7497bad0429cfe3'
    }
    # while total_num == 0 or total_num < count:
    while page<=1:
        try:
            logger.info(f'开始获取榜单数据, country: {country_id}, brand_id: {brand_id}, page: {page}')
            start_time = time.time()
            url = f"{rank_url}?market_id={market_id}&genre_id={genre_id}&country_id={country_id}&device_id={device_id}&page={page}&time={today_time_stamp}&rank_type={rank_type}&brand_id={brand_id}&k={key}"
            response = requests.get(url=url, headers=headers, proxies=PROXY).json()
            logger.success(f'获取榜单数据完成, country: {country_id}, brand_id: {brand_id}, page: {page}, 耗时: {time.time()-start_time}')
            if response['code'] == 0 and response.get('data'):
                apps = response['data']['apps']
                ranks = response['data']['ranks']
                for app in apps:
                    result = {}
                    result['id'] = app['id']
                    result['market_id'] = app['market_id']
                    result['app_id'] = app['app_id']
                    result['name'] = app['name']
                    result['logo'] = app['logo']
                    result['developer_id'] = app['developer_id']
                    result['release_time'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(app['release_time'] / 1000))
                    result['genre_id'] = app['genre_id']
                    result['genre_name'] = app['genres'][0]['name']
                    result['bundle_id'] = app['bundle_id']
                    result['last_release_time'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(app['last_release_time'] / 1000))
                    result['version'] = app['version']
                    result['title'] = app['title']
                    result['country_id'] = app['country_id']
                    result['rating'] = app['rating']
                    result['rating_count'] = app['rating_count']
                    result['price'] = app['price']
                    result['price_unit'] = app['price_unit']
                    result['category_id'] = app['category_id']
                    result['status_id'] = app['status_id']
                    result['download_times'] = app['download_times']
                    result['sizes'] = app['sizes']
                    result['mer_price'] = app['mer_price']
                    result['rating_count_list'] = app['rating_count_list']
                    result['top_in_apps'] = app['top_in_apps']
                    result['brand_id'] = ranks[0]['brand_id']
                    result['rank_type'] = ranks[0]['rank_type']
                    result['sub_brand_id'] = ranks[0]['sub_brand_id']
                    result['new_brand_id'] = ranks[0]['new_brand_id']
                    result['type'] = ranks[0]['apps'][apps.index(app)].get('type')
                    result['genre_ranking'] = ranks[0]['apps'][apps.index(app)].get('genre_ranking')
                    result['genre_ranking_incr'] = ranks[0]['apps'][apps.index(app)].get('genre_ranking_incr')
                    result['category_ranking'] = ranks[0]['apps'][apps.index(app)].get('category_ranking')
                    result['category_ranking_incr'] = ranks[0]['apps'][apps.index(app)].get('category_ranking_incr')
                    result['is_ad'] = ranks[0]['apps'][apps.index(app)].get('is_ad')
                    result['hegemony_days'] = ranks[0]['apps'][apps.index(app)].get('hegemony_days')
                    result['word_coverage'] = ranks[0]['apps'][apps.index(app)].get('word_coverage')
                    result["content"] = detail_app_process(result['genre_ranking'], result['bundle_id'], result['id'])
                    time.sleep(2)
                    result['colors'] = app_colors_process(result['bundle_id'])
                    app_records.append(result)
                    developer_records.append(app['developer'])
                page += 1
                total_num = total_num + len(apps)
                count = int(ranks[0]['count']) if count == 0 else count
            else:
                logger.error(f'获取榜单数据异常, code: {response["code"]}, message: {response["msg"]}')
        except Exception as e:
            logger.error(e)
    return app_records, developer_records


if __name__ == "__main__":
    config_name = "config-local.json"

    with open(config_name, 'r') as fp:
        config = json.loads(fp.read())

    rank_url = config['task']['rank_url']
    app_url = config['task']['app_url']
    market_id = config['task']['market_id']
    genre_id = config['task']['genre_id']
    country_id = config['task']['country_id']
    language_id = config['task']['language_id']
    device_id = config['task']['device_id']
    rank_type = config['task']['rank_type']
    brand_list = config['task']['brand_id']
    key = config['task']['key']

    path = os.path.join(os.path.abspath("./"), "/data")
    if not os.path.exists(path):
        os.makedirs(path)
    status_file = os.path.join(path, "status.json")

    thread_pool = ThreadPoolExecutor(max_workers=multiprocessing.cpu_count()*2)
    mysql = pymysql.connect(host=config['mysql']['host'],
                            port=config['mysql']['port'],
                            user=config['mysql']['user'],
                            password=config['mysql']['password'],
                            database=config['mysql']['database'])
    cursor = mysql.cursor()
    
    fetures = [thread_pool.submit(list_app_process, brand_id) for brand_id in brand_list]
    for feture in as_completed(fetures):
        developer_result, app_result = feture.result()
        start_time = time.time()
        logger.info(f'开始开发商数据入库, 数量: {len(developer_result)}')
        result = [(d['id'], d['name'], d['website'], d['email'], d['logo']) for d in developer_result]
        developer_sql = f"INSERT IGNORE INTO `developer`(`id`, `name`, `website`, `email`, `logo`) VALUES (%s, %s, %s, %s, %s)"
        cursor.executemany(query=developer_sql, args=result)
        mysql.commit()
        logger.success(f'开发商数据入库完成, 数量: {len(developer_result)}, 耗时: {time.time()-start_time}')

        logger.info(f'开始app数据入库, 数量: {len(app_result)}')
        start_time = time.time()
        result = [(app['id'], app['market_id'], app['brand_id'], app['app_id'], app['name'], app['logo'], app['developer_id'],
                app['genre_id'], app['genre_name'], app['bundle_id'], app['release_time'], app['last_release_time'],
                app['version'], app['title'], app['content'], app['colors'], app['country_id'], app['rating'], app['rating_count'],
                app['price'], app['price_unit'], app['category_id'], app['status_id'], app['download_times'], app['sizes'],
                app['mer_price'], app['type'], app['genre_ranking'], app['genre_ranking_incr'], app['category_ranking'],
                app['category_ranking_incr'], app['is_ad'], app['hegemony_days'], app['word_coverage'], datetime.now()
                ) for app in app_result]
        app_sql = f"INSERT IGNORE INTO `apps`(id, market_id, brand_id, app_id, `name`, logo, developer_id, genre_id, genre_name," \
                f"bundle_id, release_time, last_release_time, version, title, content, colors, country_id, rating, rating_count," \
                f"price, price_unit, category_id, status_id, download_times, sizes, mer_price, `type`, genre_ranking, genre_ranking_incr," \
                f"category_ranking, category_ranking_incr, is_ad, hegemony_days, word_coverage, created_time) VALUES " \
                f"(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        cursor.executemany(query=app_sql, args=result)
        mysql.commit()
        logger.success(f'app数据入库完成, 数量: {len(app_result)}, 耗时: {time.time() - start_time}')
        save_excel_file(app_result)
    cursor.close()
    mysql.close()
    

