import os.path
import re
import json
import time
import openpyxl
import requests
from loguru import logger
from openpyxl import Workbook
from datetime import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


COUNTRY = {
    101: "印度",
    3: "巴西"
}


def app_colors_process(driver, country_id, app_id, name):
    text = ""
    err = ["?", ",", "_", "/", "*", """,""", "<", ">", "|"]
    for i in err:
        name = name.replace(i, "")
    url = f"https://app.diandian.com/pk/11-{country_id}/appInfo?lang=19&sort={app_id}"
    try:
        logger.info(f"开始获取app信息, 国家: {COUNTRY[country_id]}, app: {app_id}")
        driver.get(url)
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, "app-name")))
    except:
        driver.refresh()
        try:
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, "app-name")))
        except:
            driver.refresh()
            time.sleep(5)
    time.sleep(5)
    logger.success(f"获取app信息完成, 国家: {COUNTRY[country_id]}, app: {app_id}")
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    color_tags = soup.find_all("div", class_="dd-flex dd-flex-space colot-detail-list")
    for tag in color_tags:
        color = tag.find("div", class_="color-show-word").get_text()
        num = tag.find("div", class_='color-right').get_text()
        text = text + color + "-" + num + ";"
    # js_button = 'document.documentElement.scrollTop=100000'
    # driver.execute_script(js_button)
    # time.sleep(2)
    # driver.find_element(by=By.CLASS_NAME, value="color-distribution-words dd-pointer").click()
    # driver.find_element(by=By.LINK_TEXT, value="展开").click()
    # screen_element = driver.find_element(by=By.CLASS_NAME, value="color-distribution-css")
    # screen_element.screenshot(f"/data/logo色板/{rank}-{name}.png")
    return text


def save_excel_file(driver, country_id, brand_id, app_result, file_name):
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
    else:
        wb = Workbook()
    sheet_name = "Sheet"
    if brand_id == 1:
        sheet_name = "免费榜"
    elif brand_id == 5:
        sheet_name = "人气飙升榜"
    ws = wb.create_sheet(sheet_name)
    logger.info(f"开始存储app数据, 国家: {COUNTRY[country_id]}, 榜单: {sheet_name}, 数量: {len(app_result)}")
    titles = ['排行榜', 'APP名称', '游戏类别', '下载量', '发布日期', '更新日期', 'GP链接', '应用简介', '简介翻译',
              '应用描述',
              '应用描述翻译', 'LOGO', 'LOGO主色调', 'GP展示图']
    for i in range(len(titles)):
        ws.cell(row=1, column=i + 1, value=titles[i])
    row = 2
    for app in app_result:
        bundle_id = re.search("/app/(.*?)/googleplay", app['app_url']).group(1)
        colors = app_colors_process(driver, country_id, bundle_id, app['name'])
        ws.cell(row=row, column=1, value=app['genre_ranking'])
        ws.cell(row=row, column=2, value=app['name'])
        ws.cell(row=row, column=4, value=app['download_times'])
        ws.cell(row=row, column=5, value=app.get('release_time'))
        ws.cell(row=row, column=6, value=app.get('last_release_time'))
        ws.cell(row=row, column=7, value=app['app_url'])
        ws.cell(row=row, column=8, value=app['title'])
        ws.cell(row=row, column=10, value=app['content'])
        ws.cell(row=row, column=12, value=app['logo'])
        ws.cell(row=row, column=13, value=colors)
        ws.cell(row=row, column=14, value="见之前文件")
        row = row + 1
    wb.save(file_name)
    logger.success(f"存储app数据完成, 国家: {COUNTRY[country_id]}, 榜单: {sheet_name}, 数量: {len(app_result)}, 文件: {file_name}")


def parse_rank_page(driver, brand_id, file_path):
    results = []
    soup = BeautifulSoup(driver.page_source, "html.parser")
    div_tags = soup.find_all("div", class_="el-row is-align-middle el-row--flex border dd-hover-row")
    for div in div_tags:
        result = {}
        column_tags = div.find_all("div", class_=re.compile("el-col el-col-*"))
        for column_tag in column_tags:
            if column_tag.find("div", class_="dd-app right-info"):
                head_tag = column_tag.find("div", class_="dd-app right-info")
                app_a_tag = head_tag.find("a", class_="logo-img")
                result['app_url'] = "https://app.diandian.com" + app_a_tag['href']
                result['name'] = head_tag.find("div", class_="show-text dd-max-ellipsis").get_text()
                result['developer_name'] = head_tag.find("p", class_="font-12 dd-desc-font-color develop-info").get_text()
                day_tag = head_tag.find("span", class_="day-tag").get_text()
                day = re.search("\d+", day_tag)
                result['hegemony_days'] = day.group() if day else None
            elif column_tag.find("div", class_="total-rank"):
                root_rank_tags = column_tag.find("div",
                                                 class_="dd-flex dd-align-center dd-flex-start position-relative")
                range_tag = root_rank_tags.find("div", class_=re.compile("range.*"))
                rank_tags = root_rank_tags.find("div", class_="total-rank")
                for rank_tag in rank_tags.find_all("div"):
                    if rank_tag.get_text() == "总榜":
                        result['brand_ranking'] = rank_tags.get_text(strip=True).replace("总榜", "")
                    elif rank_tag.get_text() == "游戏":
                        result['category_ranking'] = rank_tags.get_text(strip=True).replace("游戏", "")
                        if range_tag.find("i", class_="iconfont Dianxiajiang"):
                            result['category_ranking_incr'] = f"-{range_tag.get_text(strip=True)}"
                        elif range_tag.find("i", class_="iconfont Dianshangsheng"):
                            result['category_ranking_incr'] = f"+{range_tag.get_text(strip=True)}"
                    elif rank_tag.get_text() == "赌场游戏":
                        result['genre_ranking'] = rank_tags.get_text(strip=True).replace("赌场游戏", "")
                        if range_tag.find("i", class_="iconfont Dianxiajiang"):
                            result['genre_ranking_incr'] = f"-{range_tag.get_text(strip=True)}"
                        elif range_tag.find("i", class_="iconfont Dianshangsheng"):
                            result['genre_ranking_incr'] = f"+{range_tag.get_text(strip=True)}"
            elif column_tag.find("div", class_="dd-text-center"):
                if column_tags.index(column_tag) == 5:
                    result['word_coverage'] = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 7:
                    result['rating_count'] = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 8:
                    result['release_time'] = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
                elif column_tags.index(column_tag) == 9:
                    result['last_release_time'] = column_tag.find("div", class_="dd-text-center").get_text(strip=True)
            elif column_tag.find("div", role="slider"):
                star_tag = column_tag.find("div", role="slider")
                result['rating'] = star_tag['aria-valuenow']
        try:
            driver.find_element(by=By.LINK_TEXT, value=result['name'].strip()).click()
        except:
            a_tag = div.find("div", class_="show-text dd-max-ellipsis")
            a_href = "https://app.diandian.com"+a_tag.find("a")["href"]
            js = f'window.open("{a_href}");'
            driver.execute_script(js)
        time.sleep(3)
        driver.switch_to.window(driver.window_handles[1])
        try:
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, "el-image__inner")))
        except:
            driver.refresh()
            try:
                WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CLASS_NAME, "el-image__inner")))
            except:
                driver.refresh()
                time.sleep(3)
        time.sleep(2)
        detail_soup = BeautifulSoup(driver.page_source, "html.parser")
        name = detail_soup.find("div", class_="dd-flex dd-align-center dd-overflow-hidden").get_text(strip=True)
        err = ["?", ",", "_", "/", "*", """,""", "<", ">", "|"]
        for i in err:
            name = name.replace(i, "")
        title = detail_soup.find("div", class_="dd-desc-color title-text")
        title = title.get_text() if title else None
        result['title'] = title
        logo_div = detail_soup.find("a", class_="logo-wrap")
        if logo_div:
            logo = logo_div.find("img", class_="dd-app-logo")
            result['logo'] = logo["src"]
            start_save_logo_time = time.time()
            logger.info(f'开始下载logo, name: {result["genre_ranking"]}-{name}.jpg')
            try:
                logo_data = requests.get(url=logo["src"]).content
            except:
                logo_data = requests.get(url=logo["src"]).content
            with open(f"{file_path}/logo/{result['genre_ranking']}-{name}.jpg", "wb") as fp:
                fp.write(logo_data)
            logger.success(f'下载logo完成, name: {result["genre_ranking"]}-{name}.jpg, 耗时: {time.time() - start_save_logo_time}')
        backgroud_imgs = driver.find_elements(by=By.CLASS_NAME, value="el-image__inner")
        for i in range(len(backgroud_imgs)):
            start_save_screen_time = time.time()
            logger.info(f'开始下载应用截图, name: {result["genre_ranking"]}-{name}-{i + 1}.jpg')
            try:
                img_data = requests.get(url=backgroud_imgs[i].get_attribute("src")).content
            except:
                img_data = requests.get(url=backgroud_imgs[i].get_attribute("src")).content
            with open(f"{file_path}/背景图/{result['genre_ranking']}-{name}-{i + 1}.jpg", "wb") as fp:
                fp.write(img_data)
            logger.success(f'下载应用截图完成, name: {result["genre_ranking"]}-{name}-{i + 1}.jpg, 耗时: {time.time() - start_save_screen_time}')
        content = detail_soup.find("p", class_="main-content")
        result['content'] = content.get_text() if content else None
        download_times = None
        tags = detail_soup.find_all("div", class_="app-info-card-item")
        for tag in tags:
            if "累计下载量" in tag.get_text():
                download_times = detail_soup.find("a", class_="app-value dd-flex dd-flex-center dd-align-center dd-base-color font-16 font-500")
                download_times = download_times.get_text(strip=True)
        result['download_times'] = download_times
        time.sleep(3)
        if len(driver.window_handles) > 1:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        time.sleep(2)
        result['brand_id'] = int(brand_id)
        results.append(result)
    return results


if __name__ == "__main__":
    config_name = "config-local.json"
    
    with open(config_name, 'r') as fp:
        config = json.loads(fp.read())
    
    rank_url = config['task']['rank_url']
    app_url = config['task']['app_url']
    market_id = config['task']['market_id']
    genre_id = config['task']['genre_id']
    country_list = config['task']['country_id']
    language_id = config['task']['language_id']
    device_id = config['task']['device_id']
    rank_type = config['task']['rank_type']
    brand_list = config['task']['brand_id']
    key = config['task']['key']
    
    option = webdriver.ChromeOptions()
    option.add_experimental_option("detach", True)
    option.add_argument("user-agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.50'")
    option.add_argument("referer='https://app.diandian.com/rank/googleplay/11-2-33-24-0?time=1676908800000'")
    driver = webdriver.Chrome(chrome_options=option)
    file_date = datetime.strftime(datetime.now(), "%Y%m%d")
    today = f'{datetime.strftime(datetime.now(), "%Y-%m-%d")} 00:00:00'
    today_time_stamp = int(time.mktime(time.strptime(today, "%Y-%m-%d %H:%M:%S"))) * 1000
    for country_id in country_list:
        for brand_id in brand_list:
            tem = ""
            if brand_id == 1:
                tem = "01-免费榜"
            elif brand_id == 5:
                tem = "05-人气飙升榜"
            file_name = f"data/{country_id}-{COUNTRY[country_id]}/{COUNTRY[country_id]}-apps-{file_date}.xlsx"
            file_path = f"data/{country_id}-{COUNTRY[country_id]}/{tem}/{file_date}"
            logo_file = f"{file_path}/logo"
            backgroud_file = f"{file_path}/背景图"
            if not os.path.exists(file_path):
                os.makedirs(file_path)
            if not os.path.exists(logo_file):
                os.makedirs(logo_file)
            if not os.path.exists(backgroud_file):
                os.makedirs(backgroud_file)
            url = f"https://app.diandian.com/rank/googleplay/{market_id}-{rank_type}-{genre_id}-{country_id}-{brand_id}?time={today_time_stamp}"
            driver.get(url)
            time.sleep(5)
            try:
                element = driver.find_element(by=By.CLASS_NAME, value="el-dialog__wrapper")
                element.click()
            except:
                pass
            app_results = parse_rank_page(driver, brand_id, file_path)
            save_excel_file(driver, country_id, brand_id, app_results, file_name)
    driver.close()

        


